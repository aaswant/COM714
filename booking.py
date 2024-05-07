from tkinter import ttk
from tkinter import *
import openpyxl as xl
import cv2
from openpyxl import *
from PIL import Image, ImageTk
from openpyxl import workbook, load_workbook
from tkinter import messagebox
import subprocess

database = load_workbook('C:\\pycharm\\solnet\\solnet\\reserve.xlsx')

ws = database.active
ws = database.active
ws['A1'].value = "Pickup Location"
ws['B1'].value = "Drop Location"
ws['C1'].value = "Vehicle Selection"
ws['D1'].value = "Number of people"
ws['E1'].value = "Contact Number"
ws['F1'].value = "Email id"
ws['G1'].value = "Chosen Location"


class booking():
    def __init__(self, root):
        self.root = root
        self.root.title("Travel management system")
        self.root.geometry("1370x700+0+0")

        title = Label(self.root, text="Travel management system", bd=9, relief=GROOVE, font=("Algerian", 50, "bold"),
                      bg="black", fg="white")
        title.pack(side=TOP, fill=X)

        def excel():
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 40
            ws.column_dimensions['G'].width = 50

            ws.cell(row=1, column=1).value = "Pickup Location"
            ws.cell(row=1, column=2).value = "Drop Location"
            ws.cell(row=1, column=3).value = "Vehicle Selection"
            ws.cell(row=1, column=4).value = "Number of people"
            ws.cell(row=1, column=5).value = "Contact Number"
            ws.cell(row=1, column=6).value = "Email id"
            ws.cell(row=1, column=7).value = "Chosen Location"

        def focus1(event):
            course_field.focus_set()

        def focus2(event):
            sem_field.focus_set()

        def focus3(event):
            form_no_field.focus_set()

        def focus4(event):
            contact_no_field.focus_set()

        def focus5(event):
            email_id_field.focus_set()

        def focus6(event):
            address_field.focus_set()

        def clear():
            name_field.delete(0, END)


        def add():
            if (name_field.get() == "" and
                    course_field.get() == "" and
                    sem_field.get() == "" and
                    form_no_field.get() == "" and
                    contact_no_field.get() == "" and
                    email_id_field.get() == "" and
                    address_field.get() == ""):

                messagebox.showinfo("empty")

            else:
                current_row = ws.max_row
                current_column = ws.max_column

                ws.cell(row=current_row + 1, column=1).value = name_field.get()
                ws.cell(row=current_row + 1, column=2).value = course_field.get()
                ws.cell(row=current_row + 1, column=3).value = sem_field.get()
                ws.cell(row=current_row + 1, column=4).value = form_no_field.get()
                ws.cell(row=current_row + 1, column=5).value = contact_no_field.get()
                ws.cell(row=current_row + 1, column=6).value = email_id_field.get()
                ws.cell(row=current_row + 1, column=7).value = address_field.get()

                # save the file
                database.save('C:\\pycharm\\solnet\\solnet\\reserve.xlsx')

                name_field.focus_set()

                clear()

        # ===== frame generating
        manage_frame = Frame(self.root, bd=4, relief=RIDGE, bg="red")
        manage_frame.place(x=20, y=100, width=450, height=585)

        # ====heading
        m_title = Label(manage_frame, text="Booking", bg="black", fg="white", font=("times new roman", 30, "bold"))
        m_title.grid(row=0, columnspan=2, pady=30)

        # ====name

        name = Label(manage_frame, text="Pick up Location", bg="green", fg="white",
                     font=("times new roman", 20, "bold"))

        course = Label(manage_frame, text="Drop Location", bg="green", fg="white",
                       font=("times new roman", 20, "bold"))

        sem = Label(manage_frame, text="Vehicle Selection", bg="green", fg="white",
                    font=("times new roman", 20, "bold"))

        form_no = Label(manage_frame, text="No of people", bg="green", fg="white",
                        font=("times new roman", 20, "bold"))

        contact_no = Label(manage_frame, text="Contact No.", bg="green", fg="white",
                           font=("times new roman", 20, "bold"))

        email_id = Label(manage_frame, text="Email id", bg="green", fg="white",
                         font=("times new roman", 20, "bold"))

        address = Label(manage_frame, text="Chosen Location", bg="green", fg="white",
                        font=("times new roman", 20, "bold"))

        name.grid(row=1, column=0, pady=10, padx=20, sticky="w")
        course.grid(row=2, column=0, pady=10, padx=20, sticky="w")
        sem.grid(row=3, column=0, pady=10, padx=20, sticky="w")
        form_no.grid(row=4, column=0, pady=10, padx=20, sticky="w")
        contact_no.grid(row=5, column=0, pady=10, padx=20, sticky="w")
        email_id.grid(row=6, column=0, pady=10, padx=20, sticky="w")
        address.grid(row=7, column=0, pady=10, padx=20, sticky="w")

        name_field = Entry(manage_frame)
        course_field = Entry(manage_frame)
        sem_field = Entry(manage_frame)
        form_no_field = Entry(manage_frame)
        contact_no_field = Entry(manage_frame)
        email_id_field = Entry(manage_frame)
        address_field = Entry(manage_frame)





        name_field.bind("<Return>", focus1)

        course_field.bind("<Return>", focus2)

        sem_field.bind("<Return>", focus3)

        form_no_field.bind("<Return>", focus4)

        contact_no_field.bind("<Return>", focus5)

        email_id_field.bind("<Return>", focus6)




        name_field.grid(row=1, column=1, pady=10, padx=10, sticky="w")
        course_field.grid(row=2, column=1, pady=10, padx=10, sticky="w")
        combo_sem = ttk.Combobox(manage_frame, font=("times new roman", 13, "bold"), state='readonly')
        combo_sem['values'] = ("Plane", "Ferry", "Coach", "Taxi")
        combo_sem.grid(row=3, column=1, pady=10, padx=10)

        form_no_field.grid(row=4, column=1, pady=10, padx=10, sticky="w")
        contact_no_field.grid(row=5, column=1, pady=10, padx=10, sticky="w")
        email_id_field.grid(row=6, column=1, pady=10, padx=10, sticky="w")
        address_field.grid(row=7, column=1, pady=10, padx=10, sticky="w")

        # ====detail frame

        details_frame = Frame(self.root, bd=4, relief=RIDGE, bg="blue")
        details_frame.place(x=500, y=100, width=880, height=585)

        lbl_search = Label(details_frame, text="Solent Trips", bg="gray", fg="white",
                           font=("times new roman", 20, "bold"))
        lbl_search.grid(row=0, column=0, pady=10, padx=20, sticky="w")

        img = Image.open("2.jpg")
        img = img.resize((850, 540), Image.ANTIALIAS)
        self.photoimg = ImageTk.PhotoImage(img)

        lbl_pic = Label(details_frame, image=self.photoimg, bd=4, relief=RIDGE)
        lbl_pic.place(x=1, y=0, width=800, height=560)

        # ====button frame
        def insert():
            subprocess.call(["python", "trip_description.py"])


        btn_frame = Frame(manage_frame, bd=3, relief=RIDGE, bg="white")
        btn_frame.place(x=15, y=500, width=420)

        addbtn = Button(btn_frame, text="Submit", width=20, bg="grey", command=add).grid(row=0, column=1, padx=10,pady=10)
        updatebtn = Button(btn_frame, text="Trip", width=10, bg="gray", command = insert).grid(row=0, column=2, padx=10, pady=10)


class booking():
    pass
    root = Tk()
    obj = booking(root)
    root.mainloop()
