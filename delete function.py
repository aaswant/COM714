from tkinter import *
from openpyxl import *
from tkinter import messagebox

def delete():
    database = load_workbook('C:\\pycharm\\solnet\\solnet\\reserve.xlsx')
    ws = database.active
    n=int(inputtxt.get())
    ws.cell(row=n, column=1).value = ""
    ws.cell(row=n, column=2).value = ""
    ws.cell(row=n, column=3).value = ""
    ws.cell(row=n, column=4).value = ""
    ws.cell(row=n, column=5).value = ""
    ws.cell(row=n, column=6).value = ""
    ws.cell(row=n, column=7).value = ""
    print(ws.cell(row=n, column=1).value)
    database.save('C:\\pycharm\\solnet\\solnet\\reserve.xlsx')
    messagebox.showinfo("showinfo", "Successfully deleted")

root = Tk()
root.configure(background='gray')
root.title("Delete form")
root.geometry("400x150")
title = Label(root, text="Travel management system", bd=9, relief=GROOVE, font=("Algerian", 10, "bold"), bg="black",
              fg="white")
title.pack(side=TOP, fill=X)

inputtxt = Entry(root)
inputtxt.pack()
m_title = Label(root, text="Enter the number of row to delete", bg="black", fg="white", font=("times new roman", 10, "bold"))
m_title.pack()
b=Button(text="Delete", command=delete)
b.pack()
root.mainloop()
